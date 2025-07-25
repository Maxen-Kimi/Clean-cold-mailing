import pandas as pd
import re
import os
from collections import Counter
import unidecode
import openpyxl
import io
import sys
from contextlib import redirect_stdout
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

print(f"Pandas version: {pd.__version__}")
print(f"Openpyxl version: {openpyxl.__version__}")

# === Liste centralisée des prénoms/noms portugais/brésiliens les plus fréquents (normalisée) ===
EXCEPTIONS_COMPOSES_RAW = [
    'joão', 'josé', 'carlos', 'pedro', 'fernandez', 'fernandes', 'luiz', 'marco', 'rafael', 'lucas', 'andré', 'ricardo', 'vitor', 'marcos', 'daniel', 'thiago', 'paulo', 'antônio', 'bruno', 'matheus', 'felipe', 'fernando', 'maria', 'ana', 'fernanda', 'juliana', 'camila', 'patrícia', 'larissa', 'bianca', 'carla', 'priscila', 'renata', 'amanda', 'caroline', 'daniela', 'tatiane', 'gabriela', 'luana', 'letícia', 'natália', 'bruna', 'silva', 'santos', 'oliveira', 'souza', 'rodrigues', 'ferreira', 'almeida', 'lima', 'carvalho', 'pereira', 'gomes', 'martins', 'barbosa', 'teixeira', 'rocha', 'monteiro', 'moura', 'azevedo', 'vieira', 'ribeiro', 'costa', 'nascimento', 'batista', 'araújo', 'campos', 'farias', 'pinto', 'cavalcanti', 'fonseca', 'machado', 'moreira', 'da', 'de', 'do', 'das', 'dos'
]
EXCEPTIONS_COMPOSES = set([unidecode.unidecode(x).lower() for x in EXCEPTIONS_COMPOSES_RAW])

def find_column(df, variants):
    """
    Trouve la colonne du DataFrame correspondant à une des variantes fournies.
    Ignore la casse, les accents, les espaces, les underscores, les tirets.
    """
    import unidecode
    normalized = lambda s: unidecode.unidecode(str(s)).lower().replace(' ', '').replace('_', '').replace('-', '')
    norm_cols = {normalized(col): col for col in df.columns}
    for v in variants:
        v_norm = normalized(v)
        if v_norm in norm_cols:
            return norm_cols[v_norm]
    return None

PRENOM_VARIANTS = [
    'prénom', 'prenom', 'first name', 'firstname', 'first_name', 'givenname', 'given_name', 'forename', 'prénom(s)', 'prénoms', 'prename', 'prénom1', 'prenom1', 'prénom principal', 'prenom principal', 'prénomcontact', 'prenomcontact', 'contactprénom', 'contactprenom', 'prénom employé', 'prenom employe', 'prénom du contact', 'prenom du contact', 'prénom principal', 'prenom principal', 'prénom complet', 'prenom complet', 'prénom complet du contact', 'prenom complet du contact', 'first', 'f_name', 'fistname', 'fist name', 'prénom2', 'prenom2', 'prénom3', 'prenom3', 'prénom4', 'prenom4', 'prénom5', 'prenom5', 'prénom6', 'prenom6', 'prénom7', 'prenom7', 'prénom8', 'prenom8', 'prénom9', 'prenom9', 'prénom10', 'prenom10', 'prénom11', 'prenom11', 'prénom12', 'prenom12', 'prénom13', 'prenom13', 'prénom14', 'prenom14', 'prénom15', 'prenom15', 'prénom16', 'prenom16', 'prénom17', 'prenom17', 'prénom18', 'prenom18', 'prénom19', 'prenom19', 'prénom20', 'prenom20'
]
NOM_VARIANTS = [
    'nom', 'last name', 'lastname', 'last_name', 'surname', 'familyname', 'family_name', 'nom de famille', 'nomdefamille', 'nomcontact', 'contactnom', 'nom employé', 'nom employe', 'nom du contact', 'nom principal', 'nom complet', 'nom complet du contact', 'last', 'l_name', 'lastname1', 'nom1', 'nom2', 'nom3', 'nom4', 'nom5', 'nom6', 'nom7', 'nom8', 'nom9', 'nom10', 'nom11', 'nom12', 'nom13', 'nom14', 'nom15', 'nom16', 'nom17', 'nom18', 'nom19', 'nom20'
]
EMAIL_VARIANTS = [
    'email', 'e-mail', 'e_mail', 'courriel', 'mail', 'adresse email', 'adresse e-mail', 'adresseemail', 'adressemail', 'emailaddress', 'email address', 'email_adresse', 'emailadresse', 'mailadresse', 'mail address', 'adresse', 'adresse mail', 'adresse courriel', 'courrieladresse', 'courriel address', 'courriel_adresse', 'courrieladdress', 'courriel1', 'email1', 'mail1', 'email2', 'mail2', 'email3', 'mail3', 'email4', 'mail4', 'email5', 'mail5', 'email6', 'mail6', 'email7', 'mail7', 'email8', 'mail8', 'email9', 'mail9', 'email10', 'mail10'
]
DOMAIN_VARIANTS = [
    'domaine', 'domain', 'companywebsiteurl', 'website', 'siteweb', 'url', 'urlentreprise', 'urlsociete', 'urlorganisation',
    'site', 'siteentreprise', 'sitesociete', 'siteorganisation', 'web', 'webentreprise', 'websociete', 'weborganisation',
    'companydomain', 'company_domain', 'organisationdomain', 'societedomaine', 'entreprisedomaine', 'companyurl', 'organisationurl', 'societeurl', 'entrepriseurl',
    'domain1', 'domain2', 'domain3', 'domain4', 'domain5', 'domain6', 'domain7', 'domain8', 'domain9', 'domain10'
]
SOCIETE_VARIANTS = [
    'société', 'societe', 'company', 'entreprise', 'organisation', 'nomentreprise', 'nomsociete', 'nomorganisation',
    'raison sociale', 'companyname', 'organisationname', 'societename', 'entreprisename', 'company1', 'company2', 'company3', 'company4', 'company5', 'company6', 'company7', 'company8', 'company9', 'company10'
]

def extract_domain_from_email_or_url(value):
    """
    Extrait le domaine d'un email (après le @) ou d'une URL (ex: https://www.company.com -> company.com).
    Retourne une chaîne vide si rien n'est trouvé.
    """
    if pd.isna(value) or not isinstance(value, str) or not value.strip():
        return ''
    value = value.strip().lower()
    # Cas email
    if '@' in value and not value.startswith('http'):
        return value.split('@')[-1]
    # Cas URL
    match = re.search(r"(?:https?://)?(?:www\.)?([^/]+)", value)
    if match:
        domaine = match.group(1)
        # On retire les sous-domaines courants
        if domaine.startswith('www.'):
            domaine = domaine[4:]
        return domaine
    return ''

def clean_name(name):
    if pd.isna(name):
        return ""
    # Convertir en string si ce n'est pas déjà le cas
    name = str(name)
    # Supprimer les caractères spéciaux (mais garder les tirets)
    name = re.sub(r'[^a-zA-ZÀ-ÿ\-\s]', '', name)
    # Ne plus couper au tiret : on garde le nom complet
    # Nettoyer les espaces
    name = name.strip()
    # Vérifier la longueur minimale
    if len(name) <= 1:
        return ""
    return name.lower()

def generate_email(row, pattern, societe_col=None):
    if pd.isna(pattern):
        return ""
    
    def join_if_exception(name):
        parts = [unidecode.unidecode(p).lower() for p in str(name).strip().split()]
        if len(parts) > 1 and all(p in EXCEPTIONS_COMPOSES for p in parts):
            return ''.join(parts)
        # Utiliser uniquement le premier mot (prénom ou nom)
        first_part = str(name).strip().split()[0] if str(name).strip() else ''
        return clean_name(first_part)

    firstname = join_if_exception(row['Prénom'])
    lastname = join_if_exception(row['Nom'])
    # Utiliser la société uniquement si la colonne existe
    company = row[societe_col].lower() if societe_col and societe_col in row and pd.notna(row[societe_col]) else ''
    
    if not firstname or not lastname:
        return ""
        
    # Extraire la première lettre pour les initiales
    firstname_initial = firstname[0] if firstname else ""
    lastname_initial = lastname[0] if lastname else ""
    
    email = pattern.lower()
    # Remplacer d'abord les patterns d'initiales
    email = email.replace('firstnameinitial', firstname_initial)
    email = email.replace('lastnameinitial', lastname_initial)
    # Puis remplacer les noms complets
    email = email.replace('firstname', firstname)
    email = email.replace('lastname', lastname)
    email = email.replace('company', company)
    
    return email

def step1_extract_companies():
    print("🔹 Étape 1: Extraction des entreprises uniques")
    
    # Charger le fichier d'entrée
    df = pd.read_excel('input.xlsx')
    
    # Filtrer les lignes avec email qualification 'catch_all@pro' ou vide
    mask = (df['Email Qualification'].str.contains('catch_all@pro', na=False)) | (df['Email Qualification'].isna())
    df_filtered = df[mask]
    
    # Extraire les entreprises uniques
    companies = df_filtered['Société'].unique()
    companies = sorted([c for c in companies if pd.notna(c)])
    
    # Créer et sauvegarder le fichier des entreprises
    companies_df = pd.DataFrame({'Société': companies})
    companies_df.to_excel('companies.xlsx', index=False)
    
    print(f"✅ Fichier companies.xlsx créé avec succès ({len(companies)} entreprises)")

def capture_output(func, *args, **kwargs):
    output = io.StringIO()
    with redirect_stdout(output):
        try:
            func(*args, **kwargs)
            success = True
        except Exception as e:
            print(f"❌ Erreur: {str(e)}")
            success = False
    return output.getvalue(), success

def complete_name_from_linkedin(row):
    prenom, nom = row.get('Prénom', ''), row.get('Nom', '')
    url = row.get('URL Linkedin', '')
    
    def is_initial_or_empty(val):
        val = str(val).strip()
        return val == '' or len(val) == 1 or (len(val) == 2 and val[1] == '.')
    
    if not is_initial_or_empty(prenom) and not is_initial_or_empty(nom):
        return prenom, nom, False  # Aucun champ à compléter, on sort sans rien changer
    
    if pd.isna(url) or not isinstance(url, str) or '/in/' not in url:
        return prenom, nom, False  # Pas d'URL utilisable
    
    # Nettoyage robuste du slug
    slug = url.split('/in/')[-1].split('/')[0]
    slug = slug.replace('.', '-').replace('_', '-')
    slug = re.sub(r'[^a-zA-Z\-]', '', slug)
    slug_parts = [part for part in slug.split('-') if part.isalpha()]
    
    # Vérification stricte du nombre de mots
    found = False
    if not slug_parts:
        return prenom, nom, found
    
    # CORRECTION: Gérer les assignations correctement
    if is_initial_or_empty(prenom) and len(slug_parts) >= 1:
        prenom = slug_parts[0].capitalize()  # Premier élément = prénom
        found = True
    
    if is_initial_or_empty(nom):
        if len(slug_parts) >= 2:
            nom = slug_parts[1].capitalize()  # Deuxième élément = nom de famille
            found = True
        else:
            # Si on n'a qu'un seul élément et que le prénom était vide
            # on ne peut pas déterminer le nom de famille
            if is_initial_or_empty(row.get('Prénom', '')):
                found = False
                return prenom, nom, found
    
    # Vérification finale prénom != nom
    if prenom.lower() == nom.lower():
        found = False
    
    return prenom, nom, found

def step3_clean_and_complete(filename='input.xlsx'):
    print("🔹 Étape 3: Nettoyage et complétion des emails")
    
    # Vérifier que les fichiers nécessaires existent
    if not os.path.exists('detected_patterns.xlsx'):
        print("❌ Erreur: detected_patterns.xlsx non trouvé")
        return False
        
    if not os.path.exists(filename):
        print(f"❌ Erreur: {filename} non trouvé")
        return False
    
    try:
        # Charger les fichiers
        patterns_df = pd.read_excel('detected_patterns.xlsx')
        input_df = pd.read_excel(filename)
        input_df = input_df.reset_index(drop=True)

        # Détection ultra-robuste des colonnes critiques
        prenom_col = find_column(input_df, PRENOM_VARIANTS)
        nom_col = find_column(input_df, NOM_VARIANTS)
        email_col = find_column(input_df, EMAIL_VARIANTS)
        domain_col = find_column(input_df, DOMAIN_VARIANTS)
        societe_col = find_column(input_df, SOCIETE_VARIANTS)

        if not prenom_col or not nom_col:
            print("❌ Erreur: Le fichier doit contenir les colonnes Prénom et Nom (ou équivalents, voir documentation).")
            return False
        if not email_col:
            # Si la colonne email n'existe pas, on la crée vide
            email_col = 'Email'
            input_df[email_col] = ''
        if not domain_col and not societe_col:
            print("❌ Erreur: Le fichier doit contenir une colonne Domaine (ou équivalent) ou Société (ou équivalent).")
            return False

        # S'assurer que les colonnes 'Email' et 'Email Qualification' existent (pour la suite du script)
        if 'Email' not in input_df.columns:
            input_df['Email'] = ''
        if 'Email Qualification' not in input_df.columns:
            input_df['Email Qualification'] = ''
        
        # Créer un dictionnaire des patterns (par société si dispo, sinon None)
        if societe_col and societe_col in patterns_df.columns and societe_col in input_df.columns:
            patterns_dict = dict(zip(patterns_df[societe_col], patterns_df['Pattern']))
            input_df['Email Pattern'] = input_df[societe_col].map(patterns_dict)
        else:
            input_df['Email Pattern'] = None
        
        # Normaliser les colonnes Prénom et Nom (caractères spéciaux)
        input_df[prenom_col] = input_df[prenom_col].apply(lambda x: unidecode.unidecode(str(x)) if pd.notna(x) else x)
        input_df[nom_col] = input_df[nom_col].apply(lambda x: unidecode.unidecode(str(x)) if pd.notna(x) else x)

        # Supprimer les titres, diplômes et suffixes académiques/honorifiques dans Prénom et Nom
        TITRES_A_SUPPRIMER = [
            # Titres
            'dr', 'doctor', 'prof', 'professor',
            # Diplômes/suffixes
            'phd', 'ph.d', 'dphil', 'md', 'm.d', 'do', 'dvm', 'vmd', 'dds', 'dmd',
            'mba', 'emba', 'ms', 'msc', 'ma', 'm.a', 'bs', 'bsc', 'ba',
            # Autres (pharma)
            'rn', 'np', 'pa', 'facp', 'faha', 'frcp', 'facs', 'fesc'
        ]
        def remove_titles(text):
            if pd.isna(text):
                return text
            text = str(text)
            # On retire chaque mot de la liste, insensible à la casse, avec ou sans point
            for mot in TITRES_A_SUPPRIMER:
                # Mot seul ou entouré d'espaces, début ou fin de chaîne
                text = re.sub(rf'(?i)(?<![\w-]){mot}\.?\b', '', text)
            # Nettoyer les espaces multiples
            text = re.sub(r'\s+', ' ', text).strip()
            return text
        input_df[prenom_col] = input_df[prenom_col].apply(remove_titles)
        input_df[nom_col] = input_df[nom_col].apply(remove_titles)

        # === Complétion prénom/nom via LinkedIn (AVANT nettoyage) ===
        if 'URL Linkedin' in input_df.columns:
            if 'Email Qualification' not in input_df.columns:
                input_df['Email Qualification'] = ''
            for idx, row in input_df.iterrows():
                prenom, nom = row.get('Prénom', ''), row.get('Nom', '')
                new_prenom, new_nom, found = complete_name_from_linkedin(row)
                if found:
                    input_df.at[idx, 'Prénom'] = new_prenom
                    input_df.at[idx, 'Nom'] = new_nom
                # Si la complétion a échoué (besoin mais pas trouvé), on marque seulement
                elif not found and (str(prenom).strip() == '' or len(str(prenom).strip()) <= 2 or str(nom).strip() == '' or len(str(nom).strip()) <= 2):
                    input_df.at[idx, 'Email Qualification'] = 'LinkedIn name not found'

        # Nettoyer les noms (APRÈS complétion LinkedIn)
        input_df['Prénom'] = input_df['Prénom'].apply(clean_name)
        input_df['Nom'] = input_df['Nom'].apply(clean_name)

        # Sauvegarder les prénoms et noms complets dans des colonnes temporaires
        input_df['Prénom Complet'] = input_df['Prénom']
        input_df['Nom Complet'] = input_df['Nom']

        # Générer les emails avec les colonnes complètes
        def get_generated_email(row):
            # Extraire le domaine de Company Website URL si présent
            domain_from_url = ''
            if 'Company Website URL' in row and pd.notna(row['Company Website URL']):
                domain_from_url = extract_domain_from_email_or_url(row['Company Website URL'])
            pattern = None
            # Recherche dans detected_patterns.xlsx
            # 1. Par domaine (dans la liste concaténée si besoin)
            if domain_from_url:
                for idx, pat_row in patterns_df.iterrows():
                    domaines = str(pat_row.get('Domaine', '')).split(';')
                    domaines = [d.strip().lower() for d in domaines if d.strip()]
                    if domain_from_url.lower() in domaines:
                        pattern = pat_row['Pattern']
                        break
            # 2. Sinon, par société (seulement si la colonne existe)
            if pattern is None and societe_col and societe_col in row and pd.notna(row[societe_col]):
                for idx, pat_row in patterns_df.iterrows():
                    if societe_col in pat_row and str(row[societe_col]).strip().lower() == str(pat_row[societe_col]).strip().lower():
                        pattern = pat_row['Pattern']
                        break
            if pattern:
                # Utiliser les colonnes complètes pour la génération
                row_for_email = row.copy()
                row_for_email['Prénom'] = row['Prénom Complet']
                row_for_email['Nom'] = row['Nom Complet']
                # Utiliser la société uniquement si la colonne existe
                if societe_col and societe_col in row:
                    row_for_email['Société'] = row[societe_col]
                else:
                    row_for_email['Société'] = ''
                return generate_email(row_for_email, pattern, societe_col)
            return ''
        input_df['New Email'] = input_df.apply(get_generated_email, axis=1)

        # Après génération, ne garder que le premier prénom/nom dans la feuille principale
        input_df['Prénom'] = input_df['Prénom'].apply(lambda x: str(x).split()[0] if pd.notna(x) and str(x).strip() else x)
        input_df['Nom'] = input_df['Nom'].apply(lambda x: str(x).split()[0] if pd.notna(x) and str(x).strip() else x)

        # Supprimer les colonnes temporaires
        input_df = input_df.drop(['Prénom Complet', 'Nom Complet'], axis=1)
        
        # Sauvegarder le nombre de contacts avant suppression
        total_contacts_initial = len(input_df)

        # Supprimer les lignes où les 4 colonnes valent 0
        cols_to_check = [
            'Years in Position',
            'Months in Position',
            'Years in Company',
            'Months in Company'
        ]
        mask_zero = None
        if all(col in input_df.columns for col in cols_to_check):
            mask_zero = (input_df[cols_to_check] == 0).all(axis=1)
            input_df = input_df[~mask_zero].copy()

        # Supprimer les lignes où 'Connections' <= 50
        mask_conn = None
        if 'Connections' in input_df.columns:
            mask_conn = input_df['Connections'] <= 50
            input_df = input_df[~mask_conn].copy()

        # Supprimer les lignes où 'No Match Reasons' == 'company_unknown'
        mask_company_unknown = None
        if 'No Match Reasons' in input_df.columns:
            mask_company_unknown = input_df['No Match Reasons'] == 'company_unknown'
            input_df = input_df[~mask_company_unknown].copy()

        # Calculer le nombre de contacts supprimés
        contacts_supprimes = 0
        masks = [m for m in [mask_zero, mask_conn, mask_company_unknown] if m is not None]
        if masks:
            from functools import reduce
            import numpy as np
            mask_total = reduce(lambda a, b: a | b, masks)
            contacts_supprimes = np.sum(mask_total)

        # Définir les différents cas
        generated_mask = (input_df['New Email'] != '') & (input_df['New Email'] != input_df['Email'])
        failed_mask = (input_df['New Email'].isna()) | (input_df['New Email'] == '')

        # Mettre à jour Email qualification selon les cas, sans écraser 'nominative@pro'
        def update_qualification(row):
            if str(row['Email Qualification']) == 'nominative@pro':
                return 'nominative@pro'
            if generated_mask.loc[row.name]:
                return 'Generated'
            if failed_mask.loc[row.name]:
                return 'Not find'
            return row['Email Qualification']
        input_df['Email Qualification'] = input_df.apply(update_qualification, axis=1)
        
        # Supprimer uniquement la colonne temporaire de pattern
        if 'Email Pattern' in input_df.columns:
            input_df = input_df.drop('Email Pattern', axis=1)
        
        # Vérifier que la colonne New Email est toujours présente
        if 'New Email' not in input_df.columns:
            print("⚠️ La colonne New Email a disparu!")
            return False
            
        # Transformer les colonnes 'Prénom' et 'Nom' en Nom propre (après la complétion)
        if 'Prénom' in input_df.columns:
            input_df['Prénom'] = input_df['Prénom'].apply(lambda x: str(x).capitalize() if pd.notna(x) else x)
        if 'Nom' in input_df.columns:
            input_df['Nom'] = input_df['Nom'].apply(lambda x: str(x).capitalize() if pd.notna(x) else x)

        # Centraliser la liste d'exceptions et la fonction de normalisation
        def is_composed_and_not_exception(cell):
            if pd.isna(cell):
                return False
            parts = [unidecode.unidecode(p).lower() for p in str(cell).strip().split()]
            debug_exceptions = [p in EXCEPTIONS_COMPOSES for p in parts]
            if len(parts) < 2:
                return False
            if all(debug_exceptions):
                return False
            return True
        composed_mask = False
        composed_df = pd.DataFrame()
        if 'Prénom' in input_df.columns and 'Nom' in input_df.columns:
            prenom_mask = input_df['Prénom'].apply(is_composed_and_not_exception)
            nom_mask = input_df['Nom'].apply(is_composed_and_not_exception)
            # Correction : on déplace si prénom composé non exception OU nom composé non exception
            composed_mask = prenom_mask | nom_mask
            # Extraire les lignes composées
            composed_df = input_df[composed_mask].copy()
            # Supprimer ces lignes du principal (ils ne seront pas traités pour la génération d'emails)
            input_df = input_df[~composed_mask].copy()

        # Sauvegarder le résultat final dans deux feuilles du même fichier Excel
        columns_to_save = [col for col in [
            'Prénom', 'Nom', 'URL Linkedin', 'Société', 'Email', 'Email Qualification', 'New Email'
        ] if col in input_df.columns]
        with pd.ExcelWriter('cleaned_contacts.xlsx', engine='openpyxl') as writer:
            input_df[columns_to_save].to_excel(writer, index=False, sheet_name='Contacts')
            if not composed_df.empty:
                composed_columns_to_save = [col for col in [
                    'Prénom', 'Nom', 'URL Linkedin', 'Société', 'Email', 'Email Qualification', 'New Email'
                ] if col in composed_df.columns]
                composed_df[composed_columns_to_save].to_excel(writer, index=False, sheet_name='Composed_Names')

        # === Coloration en rouge des emails nominative@pro non conformes au pattern ===
        wb = load_workbook('cleaned_contacts.xlsx')
        ws = wb['Contacts']
        # Trouver les index des colonnes Email, Email Qualification et New Email
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        email_col_idx = header.index('Email') + 1
        qualif_col_idx = header.index('Email Qualification') + 1
        new_email_col_idx = header.index('New Email') + 1
        rouge = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for row in ws.iter_rows(min_row=2):
            email = str(row[email_col_idx-1].value) if row[email_col_idx-1].value is not None else ''
            qualif = str(row[qualif_col_idx-1].value) if row[qualif_col_idx-1].value is not None else ''
            new_email = str(row[new_email_col_idx-1].value) if row[new_email_col_idx-1].value is not None else ''
            if qualif == 'nominative@pro' and new_email and email.lower() != new_email.lower():
                row[email_col_idx-1].fill = rouge
        wb.save('cleaned_contacts.xlsx')

        # Calculer et afficher les statistiques détaillées
        total_generated = (input_df['Email Qualification'] == 'Generated').sum()
        total_not_find = (input_df['Email Qualification'] == 'Not find').sum()
        total_composed = len(composed_df)
        print("\n📊 Statistiques des emails générés :")
        print(f"✅ Generated : {total_generated}")
        print(f"❌ Not find : {total_not_find}")
        print(f"📝 Noms composés : {total_composed}")
        print(f"🗑️ Contacts supprimés : {contacts_supprimes}")
        print(f"📝 Total traité : {total_generated + total_not_find + contacts_supprimes}")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors du nettoyage: {str(e)}")
        return False

def analyze_email_patterns(filename=None):
    print("🔹 Analyse des patterns d'emails")
    
    if not filename:
        # Fallback pour l'utilisation en script indépendant si nécessaire
        filename = input("Entrez le nom du fichier Excel à analyser (ex: contacts.xlsx): ")
    
    if not os.path.exists(filename):
        print(f"❌ Erreur: {filename} non trouvé")
        return False
    
    try:
        df = pd.read_excel(filename)
        import unidecode
        # --- NOUVEAU : Si le fichier contient déjà Pattern et Domaine, on importe directement ---
        if 'pattern' in df.columns and 'domaine' in df.columns:
            print("📥 Import direct de patterns depuis le fichier fourni...")
            # On normalise les colonnes
            df['domaine'] = df['domaine'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else x)
            df['pattern'] = df['pattern'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else x)
            # On compte les occurrences de chaque (domaine, pattern) dans le fichier importé
            imported_counts = df.groupby(['domaine', 'pattern']).size().reset_index(name='Count')
            # On charge l'existant si présent
            if os.path.exists('detected_patterns.xlsx'):
                existing = pd.read_excel('detected_patterns.xlsx')
                # Harmonisation des colonnes existantes (si majuscules)
                if 'Domaine' in existing.columns:
                    existing = existing.rename(columns={'Domaine': 'domaine'})
                if 'Pattern' in existing.columns:
                    existing = existing.rename(columns={'Pattern': 'pattern'})
                if 'domaine' in existing.columns and 'pattern' in existing.columns:
                    existing['domaine'] = existing['domaine'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else x)
                    existing['pattern'] = existing['pattern'].apply(lambda x: str(x).strip().lower() if pd.notna(x) else x)
                    existing['Count'] = pd.to_numeric(existing['Count'], errors='coerce').fillna(1).astype(int)
                else:
                    existing = pd.DataFrame(columns=['domaine', 'pattern', 'Pourcentage', 'Count', 'Total'])
            else:
                existing = pd.DataFrame(columns=['domaine', 'pattern', 'Pourcentage', 'Count', 'Total'])
            # Fusionner intelligemment : incrémenter les counts si (domaine, pattern) existe déjà
            merged = pd.concat([
                existing[['domaine', 'pattern', 'Count']],
                imported_counts[['domaine', 'pattern', 'Count']]
            ], ignore_index=True)
            merged = merged.groupby(['domaine', 'pattern'], as_index=False)['Count'].sum()
            # Recalculer le total par domaine
            merged['Total'] = merged.groupby('domaine')['Count'].transform('sum')
            merged['Count'] = pd.to_numeric(merged['Count'], errors='coerce').fillna(1).astype(int)
            merged['Total'] = pd.to_numeric(merged['Total'], errors='coerce').fillna(1).astype(int)
            merged['Pourcentage'] = (100 * merged['Count'] / merged['Total']).round(1)
            merged = merged.sort_values(['domaine', 'Count'], ascending=[True, False])
            # Sauvegarder TOUS les patterns pour chaque domaine
            merged[['domaine', 'pattern', 'Pourcentage', 'Count', 'Total']].to_excel('detected_patterns.xlsx', index=False)
            print(f"✅ Patterns importés et fusionnés avec la base existante. {len(imported_counts)} patterns ajoutés.")
            return True
        # ... suite : logique actuelle ...
        # 1. Normalisation des noms de colonnes
        def normalize_col(col):
            return unidecode.unidecode(str(col)).lower().replace('-', '').replace(' ', '').replace('_', '')
        df.columns = [normalize_col(col) for col in df.columns]
        # 2. Dictionnaire des variantes
        col_variants = {
            'email': ['email', 'e-mail', 'e_mail', 'courriel', 'mail', 'adresseemail', 'adressemail'],
            'prenom': ['prenom', 'prénom', 'first_name', 'firstname'],
            'nom': ['nom', 'last_name', 'lastname', 'surname'],
            'domaine': [
                'domaine', 'domain', 'companywebsiteurl', 'website', 'siteweb', 'url', 'urlentreprise', 'urlsociete', 'urlorganisation',
                'site', 'siteentreprise', 'sitesociete', 'siteorganisation', 'web', 'webentreprise', 'websociete', 'weborganisation',
                'companydomain', 'company_domain', 'organisationdomain', 'societedomaine', 'entreprisedomaine', 'companyurl', 'organisationurl', 'societeurl', 'entrepriseurl'
            ],
            'societe': ['societe', 'société', 'company', 'entreprise', 'organisation']
        }
        # 3. Mapping automatique
        col_map = {}
        for key, variants in col_variants.items():
            for v in variants:
                if v in df.columns:
                    col_map[key] = v
                    break
        # Nouvelle logique : Email, Prénom, Nom, et au moins un identifiant d'entreprise
        possible_keys = [col_map.get('domaine'), col_map.get('societe')]
        required_columns = [col_map.get('email'), col_map.get('prenom'), col_map.get('nom')]
        if not all(col is not None for col in required_columns) or not any(col is not None for col in possible_keys):
            print("❌ Erreur: Le fichier doit contenir les colonnes: Email, Prénom, Nom, et au moins une colonne parmi Domaine, Company Website URL, Société")
            return False
        # On retire les lignes où Email, Prénom, Nom ou aucune clé d'entreprise n'est présente
        def has_company_key(row):
            if col_map.get('domaine') and pd.notna(row.get(col_map['domaine'], '')) and str(row.get(col_map['domaine'], '')).strip():
                return True
            if 'companywebsiteurl' in df.columns and pd.notna(row.get('companywebsiteurl', '')) and str(row.get('companywebsiteurl', '')).strip():
                return True
            if col_map.get('societe') and pd.notna(row.get(col_map['societe'], '')) and str(row.get(col_map['societe'], '')).strip():
                return True
            return False
        df = df.dropna(subset=required_columns)
        df = df[df.apply(has_company_key, axis=1)]
        patterns = []
        new_companies = set()  # Pour suivre les nouvelles clés d'entreprise
        # Préparer un mapping clé d'entreprise -> domaine site web (si colonne présente)
        website_domains = {}
        if 'companywebsiteurl' in df.columns:
            for idx, row in df.iterrows():
                # Détermination de la clé d'entreprise (priorité Domaine > URL > Société)
                if col_map.get('domaine') and pd.notna(row.get(col_map['domaine'], '')) and str(row.get(col_map['domaine'], '')).strip():
                    key = str(row[col_map['domaine']]).strip().lower()
                elif 'companywebsiteurl' in df.columns and pd.notna(row.get('companywebsiteurl', '')) and str(row.get('companywebsiteurl', '')).strip():
                    key = extract_domain_from_email_or_url(row['companywebsiteurl'])
                elif col_map.get('societe') and pd.notna(row.get(col_map['societe'], '')) and str(row.get(col_map['societe'], '')).strip():
                    key = str(row[col_map['societe']]).strip().lower()
                else:
                    continue
                url = row.get('companywebsiteurl', '')
                dom = extract_domain_from_email_or_url(url)
                if key and dom:
                    website_domains[key] = dom
        entreprises_traitees = set()
        # Déterminer si on doit filtrer sur la qualification
        filter_on_qualification = 'emailqualification' in df.columns
        # --- Compteurs pour le résumé global ---
        total_lignes = 0
        lignes_ignores = 0
        patterns_non_reconnus = 0
        domaines_ambigu = 0
        nouveaux_domaines = set()
        # --- Nouvelle logique : historique des patterns par domaine avec pourcentage ---
        # Charger l'historique si existant
        pattern_counter = defaultdict(lambda: defaultdict(int))  # {domaine: {pattern: count}}
        existing_patterns_df = None
        if os.path.exists('detected_patterns.xlsx'):
            existing_patterns_df = pd.read_excel('detected_patterns.xlsx')
        # Compter les patterns du fichier courant
        for index, row in df.iterrows():
            try:
                if not isinstance(row[col_map['email']], str):
                    lignes_ignores += 1
                    continue
                email = str(row[col_map['email']]).lower().strip()
                firstname = clean_name(row[col_map['prenom']])
                lastname = clean_name(row[col_map['nom']])
                # Détermination du domaine (clé unique)
                if col_map.get('domaine') and pd.notna(row.get(col_map['domaine'], '')) and str(row.get(col_map['domaine'], '')).strip():
                    vrai_domaine = str(row[col_map['domaine']]).strip().lower()
                elif 'companywebsiteurl' in df.columns and pd.notna(row.get('companywebsiteurl', '')) and str(row.get('companywebsiteurl', '')).strip():
                    vrai_domaine = extract_domain_from_email_or_url(row['companywebsiteurl'])
                else:
                    continue
                if not email or not firstname or not lastname or not vrai_domaine:
                    continue
                if '@' not in email:
                    continue
                if filter_on_qualification:
                    if not (("nominative@pro" in str(row.get('emailqualification', ''))) or ("generated" in str(row.get('emailqualification', '')))):
                        continue
                local_part = email.split('@')[0]
                domain = email.split('@')[1]
                firstname_initial = firstname[0] if firstname else ''
                lastname_initial = lastname[0] if lastname else ''
                pattern = local_part
                pattern_found = False
                patterns_to_try = [
                    (f"{firstname}.{lastname}", "firstname.lastname"),
                    (f"{firstname_initial}.{lastname}", "firstnameinitial.lastname"),
                    (f"{firstname}.{lastname_initial}", "firstname.lastnameinitial"),
                    (f"{firstname_initial}.{lastname_initial}", "firstnameinitial.lastnameinitial"),
                    (f"{lastname}.{firstname}", "lastname.firstname"),
                    (f"{lastname}.{firstname_initial}", "lastname.firstnameinitial"),
                    (f"{lastname_initial}.{firstname}", "lastnameinitial.firstname"),
                    (f"{lastname_initial}.{firstname_initial}", "lastnameinitial.firstnameinitial"),
                    (f"{firstname}{lastname}", "firstnamelastname"),
                    (f"{firstname_initial}{lastname}", "firstnameinitiallastname"),
                    (f"{firstname}{lastname_initial}", "firstnamelastnameinitial"),
                    (f"{firstname_initial}{lastname_initial}", "firstnameinitiallastnameinitial"),
                    (f"{lastname}{firstname}", "lastnamefirstname"),
                    (f"{lastname_initial}{firstname}", "lastnameinitialfirstname"),
                    (f"{lastname}{firstname_initial}", "lastnamefirstnameinitial"),
                    (f"{lastname_initial}{firstname_initial}", "lastnameinitialfirstnameinitial"),
                    (f"{firstname}.{lastname}.{firstname_initial}{lastname_initial}", "firstname.lastname.firstnameinitiallastnameinitial"),
                    (firstname, "firstname"),
                    (lastname, "lastname"),
                    (f"{firstname_initial}", "firstnameinitial"),
                    (f"{lastname_initial}", "lastnameinitial")
                ]
                for old, new in patterns_to_try:
                    if old == local_part:
                        pattern = new
                        pattern_found = True
                        break
                if not pattern_found:
                    continue
                full_pattern = f"{pattern}@{vrai_domaine}"
                patterns.append({
                    'domaine': vrai_domaine,
                    'pattern': full_pattern
                })
                entreprises_traitees.add(vrai_domaine)
                new_companies.add(vrai_domaine)
            except Exception as e:
                continue
        if not patterns:
            print("❌ Aucun pattern valide n'a été trouvé")
            return False
        output_file = 'detected_patterns.xlsx'
        # --- Nouvelle logique : historique des patterns par domaine avec pourcentage ---
        # Charger l'historique si existant
        pattern_counter = defaultdict(lambda: defaultdict(int))  # {domaine: {pattern: count}}
        if existing_patterns_df is not None and 'domaine' in existing_patterns_df.columns and 'pattern' in existing_patterns_df.columns and 'Count' in existing_patterns_df.columns:
            for _, row in existing_patterns_df.iterrows():
                dom = str(row['domaine']).strip().lower()
                pat = str(row['pattern']).strip().lower()
                count = int(row['Count']) if 'Count' in row and not pd.isna(row['Count']) else 1
                pattern_counter[dom][pat] += count
        # Compter les patterns du fichier courant
        for p in patterns:
            dom = str(p['domaine']).strip().lower()
            pat = str(p['pattern']).strip().lower()
            pattern_counter[dom][pat] += 1
        # Calculer le pourcentage et préparer le DataFrame final
        rows = []
        for dom, pat_dict in pattern_counter.items():
            total = sum(pat_dict.values())
            for pat, count in pat_dict.items():
                pourcentage = round(100 * count / total, 1) if total > 0 else 0
                rows.append({
                    'domaine': dom,
                    'pattern': pat,
                    'Pourcentage': pourcentage,
                    'Count': count,
                    'Total': total
                })
        patterns_df = pd.DataFrame(rows)
        # Pour la génération : ne garder qu'un seul pattern par domaine (le plus fréquent)
        patterns_df = patterns_df.sort_values(['domaine', 'Count'], ascending=[True, False])
        patterns_df = patterns_df.drop_duplicates(subset=['domaine'], keep='first')
        # Avertissement si un domaine a plusieurs patterns proches (<80% pour le majoritaire)
        for dom, group in pd.DataFrame(rows).groupby('domaine'):
            if len(group) > 1:
                max_pct = group['Pourcentage'].max()
                if max_pct < 80:
                    domaines_ambigu += 1
        patterns_df.to_excel(output_file, index=False)
        # --- Résumé global ---
        print("\n===== RÉSUMÉ DE L'ANALYSE =====")
        print(f"Total de lignes analysées : {total_lignes}")
        print(f"Total de domaines uniques : {len(pattern_counter)}")
        print(f"Nouveaux domaines ajoutés : {len(new_companies)}")
        print(f"Patterns non reconnus : {patterns_non_reconnus}")
        print(f"Lignes ignorées (email invalide ou données manquantes) : {lignes_ignores}")
        print(f"Domaines à ambiguïté (<80%) : {domaines_ambigu}")
        print(f"Total de patterns détectés (tous domaines) : {len(rows)}")
        print("==============================\n")
        if existing_patterns_df is not None:
            existing_companies = set(existing_patterns_df['domaine'])
        else:
            existing_companies = set()
        newly_added_companies = new_companies - existing_companies
        print(f"✅ Patterns détectés et sauvegardés dans {output_file}")
        print(f"   {len(patterns_df)} patterns uniques au total")
        if newly_added_companies:
            print("\n📋 Nouveaux domaines ajoutés :")
            for company in sorted(newly_added_companies):
                print(f"   • {company}")
        else:
            print("\nℹ️ Aucun nouveau domaine n'a été ajouté")
        return True
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {str(e)}")
        return False

def detect_email_pattern(emails, first_names, last_names):
    patterns = []
    for email, firstname, lastname in zip(emails, first_names, last_names):
        if pd.isna(email) or pd.isna(firstname) or pd.isna(lastname):
            continue
            
        email = email.lower()
        firstname = clean_name(firstname)
        lastname = clean_name(lastname)
        
        if not firstname or not lastname:
            continue
            
        firstname_initial = firstname[0]
        lastname_initial = lastname[0]
        
        local_part = email.split('@')[0]
        domain = email.split('@')[1]
        
        patterns_to_try = [
            # Avec des points (.)
            (f"{firstname}.{lastname}", "firstname.lastname"),
            (f"{firstname_initial}.{lastname}", "firstnameinitial.lastname"),
            (f"{firstname}.{lastname_initial}", "firstname.lastnameinitial"),
            (f"{firstname_initial}.{lastname_initial}", "firstnameinitial.lastnameinitial"),
            (f"{lastname}.{firstname}", "lastname.firstname"),
            (f"{lastname}.{firstname_initial}", "lastname.firstnameinitial"),
            (f"{lastname_initial}.{firstname}", "lastnameinitial.firstname"),
            (f"{lastname_initial}.{firstname_initial}", "lastnameinitial.firstnameinitial"),
            
            # Sans séparateurs (concaténation directe)
            (f"{firstname}{lastname}", "firstnamelastname"),
            (f"{firstname_initial}{lastname}", "firstnameinitiallastname"),
            (f"{firstname}{lastname_initial}", "firstnamelastnameinitial"),
            (f"{firstname_initial}{lastname_initial}", "firstnameinitiallastnameinitial"),
            (f"{lastname}{firstname}", "lastnamefirstname"),
            (f"{lastname_initial}{firstname}", "lastnameinitialfirstname"),
            (f"{lastname}{firstname_initial}", "lastnamefirstnameinitial"),
            (f"{lastname_initial}{firstname_initial}", "lastnameinitialfirstnameinitial"),
            
            # Pattern combiné avec points
            (f"{firstname}.{lastname}.{firstname_initial}{lastname_initial}", "firstname.lastname.firstnameinitiallastnameinitial"),
            
            # Patterns individuels (de base)
            (firstname, "firstname"),
            (lastname, "lastname"),
            (f"{firstname_initial}", "firstnameinitial"),
            (f"{lastname_initial}", "lastnameinitial")
        ]
        pattern_found = False
        for test_pattern, pattern_name in patterns_to_try:
            if test_pattern == local_part:
                patterns.append(f"{pattern_name}@{domain}")
                pattern_found = True
                break
                
        # End of Selection
    # Retourner le pattern le plus fréquent
    if patterns:
        return Counter(patterns).most_common(1)[0][0]
    return ''

def normalize_special_characters(filename=None, output_filename_web='normalized_contacts.xlsx'):
    print("🔹 Normalisation des caractères spéciaux :")
    
    is_web_call = filename is not None
    
    if not is_web_call:
        # Fallback pour l'utilisation en script indépendant si nécessaire
        filename = input("Entrez le nom du fichier Excel : ")
    
    if not os.path.exists(filename):
        print(f"❌ Erreur: {filename} non trouvé")
        return False
        
    try:
        # Charger le fichier
        df = pd.read_excel(filename)
        
    
        
        # Normaliser les colonnes First name et Last name
        df['Prénom'] = df['Prénom'].apply(lambda x: unidecode.unidecode(str(x)) if pd.notna(x) else x)
        df['Nom'] = df['Nom'].apply(lambda x: unidecode.unidecode(str(x)) if pd.notna(x) else x)
        
        # Déterminer le nom du fichier de sortie
        if is_web_call:
            final_output_filename = output_filename_web
        else:
            final_output_filename = f"{os.path.splitext(filename)[0]}_normalized.xlsx"

        # Définir les colonnes à garder dans le fichier de sortie
        columns_to_save = [col for col in [
            'Prénom', 'Nom', 'URL Linkedin', 'Société', 'Email', 'Email Qualification', 'New Email'
        ] if col in df.columns]
        # Sauvegarder le résultat
        df[columns_to_save].to_excel(final_output_filename, index=False)
        print(f"✅ Caractères spéciaux normalisés.")
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de la normalisation: {str(e)}")
        return False

def main():
    while True:
        print("\n📋 Menu Principal:")
        print("1. Extraire les entreprises uniques")
        print("2. Analyser les patterns d'emails")
        print("3. Compléter les emails et recevoir le fichier prêt")
        print("4. Normaliser les caractères spéciaux")
        print("5. Quitter")
        
        choice = input("\nChoisissez une étape (1-5): ")
        
        if choice == '1':
            step1_extract_companies()
        elif choice == '2':
            analyze_email_patterns()
        elif choice == '3':
            step3_clean_and_complete()
        elif choice == '4':
            normalize_special_characters()
        elif choice == '5':
            print("Au revoir!")
            break
        else:
            print("❌ Choix invalide. Veuillez réessayer.")

if __name__ == "__main__":
    main()